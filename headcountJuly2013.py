#!python
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import time

#I created a smaller subset file, 'headcounttest.xlsx', to work around the read/write speed issues during testing
##will this path work?: cd 'C:\Users\kidrissa\Documents\Monthly Headcount Schedule\July 2013 Headcount'
    ##Yes, but only from the shell. I'll need to bring in OS Module to make this work in code

#starting timer; basic performance profiling
start0 = time.time() #Start Loading Timer

wb = load_workbook(filename = r'July 2013 Kronos Headcount Report_working.xlsx')
source = wb.get_sheet_by_name('raw data')
end = time.time() #End Timer

durLoad = end - start0 #duration to Load
print durLoad

#Building a list of lists; each internal list represents a row of data
start = time.time() #Start Table creation timer

table = []

for row in range(len(source.rows)):
    r =[]
    #adding 'ref' selection tuple dropped loop time on 89 cols from 0.416000 sec to 0.008000
    #reordering the tuple indexs here to avoid having to shuffle them later
    #This will be less fragile if I take the following advice from Glen:
        #do this by column header/name instead of index
        #include code that will throw a VISIBLE exception if a needed column is missing
    ref = (3, 2, 0, 10, 4, 17, 21) 
    for col in ref: #Original "in" argument was 'range(len(source.columns))'
        r.append(source.cell(row = row, column = col).value)
    table.append(r)
end = time.time() #End Timer
durTable = end - start #with 1,000 rows, durLoop = datetime.timedelta(0, 47, 16000) or 47 seconds

#Only for viewing profiling results
print "Loading time for", source, ": ", durLoad
print "Time to create 'Table' from ",source, "for", len(source.rows), "rows and ", len(source.columns), "columns: ", durTable


#Getting the new internal indexes
'''
for col in range(len(table[0])):
    print table[0][col], col
''' 

'''
for i in a:
    print i
    print "\n"
'''    

###Not sure WHAT I was trying to do here, with this 'inner comment block'
#   sTART iNNER cOMMENT bLOCK
##I was trying to visualize the data after reading it in
'''
for i in range(len(table)):
    print i,'===>', table[i], '\n'
'''
                  

'''
#Generating the Row & Column indexes
for row in range(len(table)):
    print row, "\n"
    for col in range(len(table[0])):
        print col, table[row][col]
'''        
#Creating a spreadsheet in memory; Writing results to it (in memory)
#This section will be replaced by the final report spreadsheet
#that will have the following columns
#CompNum, CC, EmpNum, EmpName, MngrName, TotalDOEhours, TotalProjhours

start = time.time() #Start timer for creating 'Target' workbook
target = Workbook()
dest_filename = r'hdcntsum.xlsx'

#regarding a question about relative paths
# both r'Downloads\hdcntsum.xlsx' and r'..\hdcntsum.xlsx worked
#haven't tested networked drives yet
#my theory is 
    #r'..\..\..\M:\Dbsteam\BUDGET\Jackie\MNTH_RPT\2013\June 2013\Headcount Misc\hdcntsum.xlsx' SHOULD work
#Maybe not: r'\hdcntsum.xlsx' saved the file to the ROOT drive; C:

##Commenting out this next section; I don't need it. It was included before as a test
##to be sure I was getting the proper data
#ws = target.worksheets[0]
#ws.title = "Monthly Headcount Raw"
#
#for row in range(len(table)):
#    for col in range(len(table[0])):
#        ws.cell(row = row, column = col).value = table[row][col]
#        
#end = time.time()  #End timer for creating 'Target' workbook
#durTarget = end - start #durATION for Target
        
#In an earlier iteration, I wrote the above file here. I've since postponed the write
#and made 'Table' into the first sheet in the workbook. It's the 'reduced' form
#of the raw data needed to create the final report. It can also be used as a check.

#Creating list of keys
#Each key is (as of 2013-07-16)a LIST made up of [Company Number, CC, EmpNum]. It SHOULD be
#a TUPLE made up of (Company Number, CC, EmpNum). But I was having trouble with Tuples
#I'll fix that in a refactor. Tuples are better for keys. Immutibilty FTW!
start = time.time() #Start timer for creating 'Keylist' workbook
keylist = []

for r in range(len(table)): #I need to change the range to (1, len(table)), to get rid of the header key
    newkey = table[r][:3] #Original code was 'newkey = tuple(table[r][:3])'
    if newkey not in keylist:
        keylist.append(newkey)
end = time.time()  #End timer for creating 'Keylist' 
durKeylist = end - start #durATION for Keylist
                

#Comparing keylist items to table rows and calculating hour totals
#using 'r' as shorthand for 'row', to avoid namespace confusion
start = time.time() #start Hourtable timer
hourtable = []
#'hourtable' will be a list of lists; each sublist is composed of a tuple and two ints; 
#the first element of each list is the composite key Index 1 is DOE; Index 2 is Project;
#DOE and Project are the hourly totals FOR that key
for key in keylist:
    doe = 0 #Counter for DOE Hours
    project = 0 #Counter for Project hours
    newrow =[]
    for r in range(len(table)):
        #newrow = []
        if key == table[r][:3]: #Tuple version: 'if key == tuple(table[r][:3])'
            if table[r][5] == 'DOE':
                doe = doe + table[r][6]
            elif table[r][5] == 'Project':
                project = project + table[r][6]

#THIS NEXT BLOCK IS A HOT MESS! It comes from me not knowing how to work with Dictionaries
#I got some good tips at PyHou on 2013-07-16 and I'll apply those in the refactor
##PRIMARY DICTIONARY PROBLEM: 
## for some reason I was thinking dictionaries could only hold a SINGLE {key: value} pair
    #original statement was "hourtable.append({key : (doe, project)})"
    #hourtable.append({key : [doe, project]})
    newrow.extend(key) #add key values to the new row
    #newrow.extend(table[r][3:5]) #add employee & manager names to new row
    newrow.extend([doe, project]) #add total DOE & Project hours to new row
    hourtable.append(newrow) #add new row to table; Was hourtable.append([key, doe, project])
    #newrow = (table[r][:5])
    #hours = [doe, project] 
    #newrow.extend(hours)
    #hourtable.append(table[r][:5])
end = time.time()  #End timer for creating 'Hourlist'
durHourlist = end - start

'''
#I'm going to add the Unicode conversion code here
unicode_rows = [u'Company', u'CC', u'Employee Num', u'Employee Name', u'Manager']
for ur in table: #namespace issues with "r"; using 'ur' for 'Unicode Row' here
    ri = table.index(ur) #THIS lets me avoid that 'range(len(nestedTable))' nonsense
    for header in unicode_rows:
        table[ri][unicode_rows.index(header)] = unicode(table[ri][unicode_rows.index(header)])
#End of Unicode conversion code; 
'''
                
#Create final output table
start = time.time() #start finaltable creation timer

finaltable =[]
for k in range(len(keylist)):
    finalrow = []
    for r in range(len(table)):
        if keylist[k][:3] == table[r][:3]: 
            finalrow = table[r][:5] + hourtable[k][-2:]
    finaltable.append(finalrow)

end = time.time()  #End 'finaltable' timer
durFinalTable = end - start

#Set all columns except DOE & Project Hours to Unicode strings
'''
unicode_rows = [u'Company', u'CC', u'Employee Num', u'Employee Name', u'Manager']    
for r in finaltable:
    ri = finaltable.index(r) #THIS lets me avoid that 'range(len(nestedTable))' nonsense
    for header in unicode_rows:
        finaltable[ri][unicode_rows.index(header)] = unicode(finaltable[ri][unicode_rows.index(header)])

#Are the values Unicode strings before they're written to the cells?
print finaltable[:5]
'''

#Write final workbook to memory and save to file

start = time.time() #start Target final spreadsheet write to memory timer

ws1 = target.create_sheet(0)
ws1.title = "Monthly Headcount Summary"
for row in finaltable:
    rowIn = finaltable.index(row)
    for col in range(len(finaltable[0])): #changed from "range(len(finaltable[0]))"
        #colIn = row.index(col) #colIn replaces col as the Column Indexes below
        ws1.cell(row = rowIn, column = col).value = finaltable[rowIn][col]
        
end = time.time() #End Target final spreadsheet write to memory timer
durFinalTableMem = end - start

###I tried to move the 'create_tabs' function into this module. It's not working
###So, I'm going to create a "totalTable" in the 'functionalSheets' module
##creating the header row
#header = [] 
#for c in finaltable[0]:
#    header.append(c)
##I STILL don't know why those 0s got added to the header; need to figure that out
##Probably because the source data had ONE 'DOE/Project' column, not two separate cols
#header[-2:] = 'DOE', 'Project'
#
#
###I'm using the function that creates the individual functional area worksheets 'functionalSheets.py'
###to create the 'neat, sorted' version of the monthly summary
#def create_tabs(functable, tabname):
#
#    '''
#    list of lists, string --> list of lists
#    
#    Takes in nested list for each functional area (created by funcTable/makeSubseaTable/makeNoSubseaTable) and a string (tabname)
#    each inner list reperesents a row of data; creates a spreadsheet in memory, writes those rows to the spreadsheet
#    The string becomes the name of the worksheet
#    '''
#
#    #creating & naming the spreadsheet in memory
#    ws = target.create_sheet(0)
#    ws.title = tabname
#    #ws.append(header) THIS was a mistake.
#    
#    
#    #sorted_by_second sorts the incoming data by the 2nd element in each list; the Cost Center in this case
#    #thanks StackOverflow! http://stackoverflow.com/questions/3121979/how-to-sort-list-tuple-of-lists-tuples
#    #sorted_by_second = sorted(functable, key=lambda tup: tup[1])
#    
#    ##I upgraded to using the Operator Module, as it lets me sort by MULTIPLE criteria; 
#    ##http://docs.python.org/2/howto/sorting.html#operator-module-functions
#    from operator import itemgetter
#    ##using Operator module, sorting by Cost Center, then Company, then Emp. Name
#    headcount_sorted = sorted(functable, key = itemgetter(1, 0, 3))
#    
#    #goes through the sorted nested list, writing it to the spreadsheet in memory
#    #Updated version uses the APPEND method from http://pythonhosted.org/openpyxl/api.html#module-openpyxl-worksheet-worksheet
#    
#    #spacer added to create break for manual insertion of Cost Center sum functions
#    spacer = [None, None, None, None, None, None, None]
#    for r in headcount_sorted:
#        ri = headcount_sorted.index(r)
#        #If this is the FIRST row, append the header, then the row
#        if ri == 0:
#            ws.append(header)
#            ws.append(r)
#        #If the Cost Center in THIS row is the same as the one before it, append the row
#        elif headcount_sorted[ri][1] == headcount_sorted[(ri-1)][1]:
#            ws.append(r)
#        #If this Cost Center is different than the prior row, it's a new Cost Center
#        #insert two spacers, write the header and append the row
#        else:
#            ws.append(spacer) #One for summation of the section above
#            ws.append(spacer) #One for readability
#            ws.append(header)
#            ws.append(r)
#
#
#create_tabs(finaltable, 'Monthly Headcount Sorted') #Got this error: "SheetTitleException: Maximum 31 characters allowed in sheet title"


#Finding out what data types are in the various columns
#after I figured out how to get the indexes of the items I was finding the type for "finaltable[1].index(item)",
#I plugged that in to get the header description from Row 0 of 'finaltable'
'''
for item in finaltable[1]:
    print finaltable[0][finaltable[1].index(item)], item, type(item), finaltable[1].index(item)
''' 

    
                 
#Writing that worksheet to a file
start = time.time() #start Target final spreadsheet write to file timer

target.save(dest_filename)

end0 = time.time() #End Target final spreadsheet write to file timer
durFinalTableFile = end0 - start

durTotal = end0 - start0 

#Printing my timing variables
print "Loading time for", source, " :", durLoad
print len(source.rows),"Rows; ", len(source.columns), "Columns"
print "durTable", durTable
print len(table),"Rows; ", len(table[0]), "Columns"
print "durKeylist", durKeylist
print "durHourlist", durHourlist
print "durFinalTable", durFinalTable
print "durFinalTableMem", durFinalTableMem
print "Writing time for", dest_filename, " :", durFinalTableFile
print len(finaltable),"Rows; ", len(finaltable[0]), "Columns"
print "durTotal", durTotal

###The rest of the file is unused. I need to look at that code to see if there's anything I can/want to salvage
'''
for k in range(len(keylist)):
    for key in keylist:
        print hourtable[k]
'''
#Pulling the hours out of the 'hours' list of dictionaries
#Need to further smooth this out so I can write the values to their proper places in the spreadsheet
'''for hr in range(len(hours[1][key])):
    print hours[1][key][hr]
        
'''                        
