#!python

##This will take the output of "headcount.py" and create the various functional spreadsheet summaries
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import time

start = time.time() #timer, until I learn to use http://docs.python.org/2/library/profile.html
wb = load_workbook('hdcntsum.xlsx')
source = wb.get_sheet_by_name('Monthly Headcount Summary')

#creating the header row
header =[] 
for c in source.rows[0]:
    header.append(c.value)
#I STILL don't know why those 0s got added to the header; need to figure that out
#Probably because the source data had ONE 'DOE/Project' column, not two separate cols
header[-2:] = 'DOE', 'Project'
header.extend(['Tot. Hours', 'DOE Util %', 'Proj. Util %']) 



#This function takes in the list of cost centers which make up a Functional Tab, returns a list of lists; 
#each internal list is a row for that Functional tab
##I'd originally intended it to return a list of tuples; why the change? Probably lack of skill on my part
##I could tuple() each temprow BEFORE I append it to tempTable; or learn how to build tuples programmatically
##The second seems the best option; one or both will have to happen in a later refactor
###As of 2013-08-01, This function is being cloned into two different functions;One that excludes Company 2231(Subsea)
###and another that excludes anything NOT Company 2231(Subsea) This is my current (2013-08-01)
###solution to the issue of Subsea having cost centers used in 2231 and 1902 in a Cost Center-focused report

#def functionTable(list):
#    
#    ''' 
#    list ==> list of lists
#    
#    Input a list of cost centers for a functional tab;
#    return a list of lists where each inner list is a row of representing one person in that Cost Center
#    '''
#    
#    tempTable = []
#    for row in source.rows:
#        ri = source.rows.index(row)
#        if str(row[1].value) in list:
#            temprow = []
#            for cell in row:
#                ci = source.rows[ri].index(cell)
#                temprow.append(source.rows[ri][ci].value)
#            tempTable.append(temprow)
#    return tempTable

##This function is intended to create a table of the ENTIRE contents of the headcount summary file
fullTable = []
time1 = time.time()

##The 'Monthly Headcount Summary' tab in 'hdcntsum.xlsx' has 0 in columns F&G (index[5]&[6] in Python) 
##in the first and last row. Until I figure out how to STOP that, I'm slicing off those rows.
##Truth be told, I don't NEED them. Well, I MIGHT need the first row for header names. To make the
##design less fragile
for row in source.rows[1:-1]: #trying to work around problem with 1st and last rows; was 'for row in source.rows'
    ri = source.rows.index(row)
    temprow = []
    for cell in row:
        ci = source.rows[ri].index(cell)
        #print source.rows[ri][ci].value
        temprow.append(source.rows[ri][ci].value)
        #print temprow #only for debugging purposes; I want to see when/where the float division by zero problem is happening
        #It seems to be in the "hdcntsum.xlsx" header row, where the DOE & Project columns have '0' in them.
        #I'm taking a slice of source.rows to ignore that first row for now
    temprow.append(temprow[-1]+temprow[-2]) #Total Hours: sum of DOE & Proj Hours
    temprow.append(temprow[-3]/float(temprow[-1])) #DOE Util%; DOE Hours / newly added Total
    temprow.append(temprow[-3]/float(temprow[-2])) #Proj Util%; Proj. Hours / Total
    fullTable.append(temprow)
time2 = time.time()
print "fullTable Creation Time was ", time2-time1, "seconds."

#This function takes in the list of cost centers in Subsea, returns a list of lists; 
#each internal list is a row for Subsea. It also excludes any rows that aren't
#Subsea (Company 2231)
def makeSubseaTable(list):
    
    ''' 
    list ==> list of lists
    
    Input a list of cost centers for a functional tab;
    return a list of lists where each inner list is a row of representing one person in that Cost Center
    '''

    tempTable = []
    for row in source.rows:
        ri = source.rows.index(row)
        if row[0].value == 2231: #this is what was missing. I need to test the VALUE of what's in that cell
            if str(row[1].value) in list: #row[1] is the position of the Cost Center; I should change that to get the index of the name
                temprow = []
                for cell in row:
                    ci = source.rows[ri].index(cell)
                    temprow.append(source.rows[ri][ci].value)
                temprow.append(temprow[-1]+temprow[-2]) #Total Hours: sum of DOE & Proj Hours
                temprow.append(temprow[-3]/float(temprow[-1])) #DOE Util%; DOE Hours / newly added Total
                temprow.append(temprow[-3]/float(temprow[-2])) #Proj Util%; Proj. Hours / Total
                tempTable.append(temprow)
    return tempTable

#This function takes in the list of cost centers from the various functional areas, returns a list of lists; 
#each internal list is a row in the table for that area. It also excludes any rows that ARE
#Subsea (Company 2231)
def makeNoSubseaTable(list):
    
    ''' 
    list ==> list of lists
    
    Input a list of cost centers for a functional tab;
    return a list of lists where each inner list is a row of representing one person in that Cost Center
    '''

    tempTable = []
    for row in source.rows:
        ri = source.rows.index(row)
        if row[0].value != 2231: #this is what was missing. I need to test the VALUE of what's in that cell
            if str(row[1].value) in list: #row[1] is the position of the Cost Center; I should change that to get the index of the name
                temprow = []
                for cell in row:
                    ci = source.rows[ri].index(cell)
                    temprow.append(source.rows[ri][ci].value)
                temprow.append(temprow[-1]+temprow[-2]) #Total Hours: sum of DOE & Proj Hours
                temprow.append(temprow[-3]/float(temprow[-1])) #DOE Util%; DOE Hours / newly added Total
                temprow.append(temprow[-3]/float(temprow[-2])) #Proj Util%; Proj. Hours / Total
                tempTable.append(temprow)
    return tempTable
    
    
##Creating the target workbook
target = Workbook()
dest_filename = r'hdcntfunc.xlsx'

##Function that creates the individual functional area worksheets IN the workbook just created
def create_tabs(functable, tabname):
    '''
    list of lists, string --> list of lists
    
    Takes in nested list for each functional area (created by funcTable/makeSubseaTable/makeNoSubseaTable) and a string (tabname)
    each inner list reperesents a row of data; creates a spreadsheet in memory, writes those rows to the spreadsheet
    The string becomes the name of the worksheet
    '''
    #creating & naming the spreadsheet in memory
    ws = target.create_sheet(0)
    ws.title = tabname
    #ws.append(header) THIS was a mistake.
    
    #sorted_by_second sorts the incoming data by the 2nd element in each list; the Cost Center in this case
    #thanks StackOverflow! http://stackoverflow.com/questions/3121979/how-to-sort-list-tuple-of-lists-tuples
    #sorted_by_second = sorted(functable, key=lambda tup: tup[1])
    
    ##I upgraded to using the Operator Module, as it lets me sort by MULTIPLE criteria; 
    ##http://docs.python.org/2/howto/sorting.html#operator-module-functions
    from operator import itemgetter
    ##using Operator module, sorting by Cost Center, then Company, then Emp. Name
    
    import sortCriteria #generates itemgetter keys based on the header values, instead of hardcoding them
    sort_by = sortCriteria.sort_criteria(source.rows[0]) 
    headcount_sorted = sorted(functable, key = itemgetter(sort_by[0], sort_by[1], sort_by[2]))
    
    #goes through the sorted nested list, writing it to the spreadsheet in memory
    #Updated version uses the APPEND method from http://pythonhosted.org/openpyxl/api.html#module-openpyxl-worksheet-worksheet
    
    #spacer added to create break for manual insertion of Cost Center sum functions
    spacer = [None for i in range(len(functable[0]))] #used range in a list comprehension to build this

    #bring in my custom Footer code and generate the Footer dictionary for this Functional area
    ##since I didn't adjust the Path variable (I'll do that later), costCenterFooter.py had to
    ##be in the same directory as the files it was working on
    ###Forgot I need to do moduleName.functionName() when calling a func from a module
    ###Since I named my module AND function the same thing (I hadn't planned on it being a module)
    ###I ended up with moduleName.moduleName() 

    import costCenterFooter
    footer = costCenterFooter.costCenterFooter(functable)

    for r in headcount_sorted:
        ri = headcount_sorted.index(r)
        #If this is the FIRST row, append the header, then the row
        if ri == 0:
            ws.append(header)
            ws.append(r)
        #If the Cost Center in THIS row is the same as the one before it, append the row
        elif headcount_sorted[ri][1] == headcount_sorted[(ri-1)][1]:
            ws.append(r)
        #If this Cost Center is different than the prior row, it's a new Cost Center
        #insert footer (for the previous Cost Center), one spacer, then write the header and append the row
        else:
            ws.append(footer[str(headcount_sorted[ri-1][1])]) #footer for the previous Cost Center
            ws.append(spacer) #spacer for readability
            ws.append(header)
            ws.append(r)
    ws.append(footer[str(headcount_sorted[ri][1])]) #footer for the FINAL Cost Center
    
    #Once all the Cost Centers are done, add in the grand totals for the Functional Area
    import deptTotal
    DeptTotals = deptTotal.deptTotal(footer, tabname)
    ws.append(spacer) #spacer for readability
    ws.append(DeptTotals[tabname]) #Functional Area Totals & Utilization
                

#My "Main Loop"; running the data through the two functions

##Function 1: functionTable ==> makeSubSeaTable/makeNoSubseaTable
time1 = time.time()
import json
dept_dict = json.load(file('costCenter_Function_map.json'))
sheet_dict = {}
for key in dept_dict.keys():
    if key != 'Subsea':
        sheet_dict.update({key : makeNoSubseaTable(dept_dict[key])})
    else:
        sheet_dict.update({key : makeSubseaTable(dept_dict[key])})

time2 = time.time()
print "Creation Time  for all Functional Tables was ", time2-time1, "seconds."

##Function 2: create_tabs
time1 = time.time()
for key in sorted(sheet_dict.keys(), reverse = True):
    create_tabs(sheet_dict[key], key)
create_tabs(fullTable, 'Headcount Summary Sorted')
time2 = time.time()

#print "Length of all tables is", len(subsTable + estSTable + prjCTable + infoTable + procTable + legaTable + engiTable + humaTable + prjMTable + accoTable + ethiTable + hsesTable + qualTable)
print "Creation Time for ALL tabs was ", time2-time1, "seconds."

#remove 'Sheet' worksheet, that gets created by default
target.remove_sheet(target.get_sheet_by_name("Sheet")) #the .remove_sheet() function seems to REQUIRE a worksheet object, not just a name

#Make headers bold; format will probably be the later home of functions for number & alignment formatting
import format
format.bold_headers_footers(target)

##Writing that worksheet to a file
##Moved later in file, so Exceptions sheet could be written


end = time.time()
dur = end - start
print "Total processing time", dur

def funcSheets_check_figures(d, ft):
    '''
    dict, list of lists --> dict of dicts

    takes the final dictionary of worksheets, shows the number of DATA rows(which equals Employees).
    Also calcuates the total number of DATA rows (which equals Employees) 'Headcount Summary Sorted' should have
    as well as the number of rows by which HSS & the total functional sheets differ.
    Returns a dict of dictionaries of {EmployeeNum: [Utilziation Data for that Employee]} to show who the exceptions are

    '''

    print "############################################"
    print "CHECK FIGURES"
    print "############################################"

    for key in d.keys():
        print key, "has", len(d[key]), "employees."

    check_sum = 0
    for key in d.keys():
        check_sum = check_sum + len(d[key])
    print "********************************************"
    print "Combined, the Functional Sheets have a total of", check_sum, "employees."
    print "This is compared to", len(ft), "employees in Headcount Summary Sorted."
    print "This means", len(ft) - check_sum, "employees from Headcount Summary aren't included in the functional sheets." 
    #I should include an IF here to change the sentence based on the difference direction

    ###This is where the Exceptions sheet gets created
    ###This code is too 'fragile', since I hard-coded index numbers for prototyping speed
    ###I need to undo that to make the code more robust. It probably doesn't need to be anti-fragile
    ###I'm not sure I'd know how to do that anyway.

    ##Get list of employee numbers from fullTable; hss stands for Headcount Summary Sorted, first tab in output
    hss_employees = []
    for i in ft:
        hss_employees.append(i[3])

    hss_set = set(hss_employees)

    ###Get list of Employee nums from sheet_dict
    #func stands for Functional Areas, what I call each Cost Center group;
    #Accounting, Structural Engineering, etc
    func_employees = []  
    for key in d.keys():
        for i in d[key]:
            func_employees.append([i][0][3])
    func_set = set(func_employees)

    ##Find the Exceptions as a difference between the sets
    exceptions = set.difference(hss_set, func_set)
    print exceptions
    len(exceptions)

    ##create a dictionary of {employeeNum:[utilization data]}, to write to spreadsheet
    exception_dict = {}
    for x in sorted(exceptions):
        for i in ft:
            if i[3] == x:
                exception_dict.update({x : i})

    ws = target.create_sheet(0)
    ws.title = "Exceptions"

    for key in exception_dict:
        ws.append(exception_dict[key])

    ##I PROBABLY don't need to return anything, since I'm writing the sheet inside the function
    ##But it does make printing the final message convenient. Immediate feedback.
    return exception_dict

problems = funcSheets_check_figures(sheet_dict, fullTable)

target.save(dest_filename)

print "There are", len(problems), "exceptions."